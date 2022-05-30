import datetime
import json

from django.core.files.base import ContentFile
from django.core.files.storage import FileSystemStorage
from openpyxl import load_workbook


def create_command(command: str, target: str = None, value: str = "", targets=None, description: str = "") -> dict:
    d = {
        "Command": command,
        "Value": value,
        "Description": description
    }
    if target:
        d['Target'] = target
    if targets:
        d['Targets'] = targets
    return d


def create_click_command(*args) -> dict:
    return create_command('click', *args)


def create_xpath_click_command(xpath, *args):
    target = 'xpath=' + xpath
    return create_click_command(target, *args)


class Commands:
    possible_modes = (
        'campaigns',
        'last_week',
        'last_month',
        'generate_reports',
        'download_reports',
    )

    def __init__(self, account: str):
        self.account = account

        self.timeout_wait = create_command('store', '90', '!TIMEOUT_WAIT')

        self.open_agency_panel = create_command('selectWindow', 'tab=open',
                                                'https://ads.allegro.pl/panel/agency/clients')
        self.open_clients_list = create_click_command('linkText=Przełącz na klienta')
        self.open_client_account = create_xpath_click_command(f"(//*[text()='{account}']) ")
        self.open_stats = create_xpath_click_command(
            '//*[@id="main"]/div[2]/div/header/div/div/div[2]/div/div/div/div/div[1]/nav/a[2]')
        self.open_graphic = create_xpath_click_command('//*[@id="layoutBody"]/div/div/div[2]/div[1]/div/div/a[2]')
        self.open_calendar = create_xpath_click_command('//*[@id="layoutBody"]/div/div/div/div[2]/div')

        # calendar range options
        self.select_yesterday = create_xpath_click_command("//*[text()='Wczoraj']")
        self.select_today = create_xpath_click_command("//*[text()='Dzisiaj']")
        self.select_last_week = create_xpath_click_command("//*[text()='Ostatnie 7 dni']")
        self.select_last_month = create_xpath_click_command("//*[text()='Ostatnie 30 dni']")
        self.select_last_billing_month = create_xpath_click_command("//*[text()='Poprzedni okres rozliczeniowy']")
        self.select_current_billing_month = create_xpath_click_command("//*[text()='Bieżący okres rozliczeniowy']")
        # updates calendar range
        self.update = create_xpath_click_command("//*[text()='Aktualizuj']")

        self.select_cost = create_command('select',
                                          "xpath=//*[@id=\"layoutBody\"]/div/div/div[3]/div/div/div/div/div/div/div[2]/select",
                                          'label=Koszt', [
                                              "xpath=//*[@id=\"layoutBody\"]/div/div/div[3]/div/div/div/div/div/div/div[2]/select",
                                              "xpath=//select",
                                              "css=#layoutBody > div > div > div:nth-child(3) > div > div > div > div._1kh9d > div.wrapper__eagO14H._1i06w._xbsje > div:nth-child(1) > div:nth-child(2) > select"
                                          ])

        self.select_return = create_command('select',
                                            "xpath=//*[@id=\"layoutBody\"]/div/div/div[3]/div/div/div/div/div/div/div[4]/select",
                                            'label=Zwrot z inwestycji', [
                                                "xpath=//*[@id=\"layoutBody\"]/div/div/div[3]/div/div/div/div/div/div/div[4]/select",
                                                "xpath=//div[4]/select",
                                                "css=#layoutBody > div > div > div:nth-child(3) > div > div > div > div._1kh9d > div.wrapper__eagO14H._1i06w._xbsje > div:nth-child(1) > div:nth-child(4) > select"
                                            ])

        # reports downloading commands
        self.select_offers_view = create_xpath_click_command(
            '//*[@id="layoutBody"]/div/div/div[4]/div[1]/div/div[3]/button')
        self.open_files = create_xpath_click_command(
            '//*[@id="main"]/div[2]/div/header/div/div/div[2]/div/div/div/div/div[1]/nav/div/a')
        self.click_generate_report = create_xpath_click_command('//*[@id="layoutBody"]/div/div/div[4]/div[2]/button[3]')
        self.click_download_file = create_xpath_click_command(
            '//*[@id="layoutBody"]/div/div[2]/div/div[2]/div/div[6]/div[2]/button[2]')

    @property
    def campaign_commands(self):
        return [
            self.open_agency_panel,
            self.open_clients_list,
            self.open_client_account,
        ]

    @property
    def last_week_commands(self):
        commands = self.campaign_commands + [
            self.open_stats,
            self.select_cost,
            self.select_return,
            self.open_calendar,
            self.select_last_week,
            self.update,
        ]
        # if self.account.is_graphic:
        #     commands = commands + self.graphic_commands

        return commands

    @property
    def last_month_commands(self):
        commands = self.campaign_commands + [
            self.open_stats,
            self.select_cost,
            self.select_return,
            self.open_calendar,
            self.select_last_month,
            self.update,

        ]

        # if self.account.is_graphic:
        #     commands = commands + self.graphic_commands

        return commands

    @property
    def graphic_commands(self):
        return self.campaign_commands + [
            self.open_stats,
            self.open_graphic,
            self.open_calendar,
            self.select_last_month,
            self.update,
        ]

    @property
    def generate_reports_commands(self):
        return self.campaign_commands + [
            self.open_stats,
            self.open_calendar,
            self.select_last_billing_month,
            self.update,
            self.select_offers_view,
            self.click_generate_report,
            self.open_client_account,
        ]

    @property
    def download_reports_commands(self):
        return self.campaign_commands + [
            self.open_files,
            self.click_download_file,
            self.open_client_account,
        ]

    def get_commands_array(self, mode):
        if mode not in self.possible_modes:
            raise ValueError(
                f'mode : "{mode}" is not allowed. Please select from one of possible modes: {self.possible_modes}')

        commands_for_modes = {
            self.possible_modes[0]: self.campaign_commands,
            self.possible_modes[1]: self.last_week_commands,
            self.possible_modes[2]: self.last_month_commands,
            self.possible_modes[3]: self.generate_reports_commands,
            self.possible_modes[4]: self.download_reports_commands,
        }
        return commands_for_modes[mode]


def generate_json_script(mode: str, accounts_list: list, dump: bool = False):
    commands_array = [{
        "Command": "store",
        "Target": "90",
        "Value": "!TIMEOUT_WAIT",
        "Description": ""
    }, ]

    for account in accounts_list:
        commands = Commands(account)
        commands_array.extend(commands.get_commands_array(mode))

    if dump:
        with open('../macro.json', 'w') as f:
            json.dump(commands_array, f, ensure_ascii=False)
    else:
        return commands_array


def generate_macro(mode: str, accounts_list: list, name: str):
    commands_array = generate_json_script(mode, accounts_list)
    today = datetime.date.today()
    full_json = {
        "Name": name,
        "CreationDate": f'{today.year}-{today.month}-{today.day}',
        "Commands": commands_array
    }
    fs = FileSystemStorage()
    return fs.save(name + '.json', ContentFile(json.dumps(full_json, ensure_ascii=False)))


def create_macros(file):
    wb = load_workbook(file)
    ws = wb.active
    accounts = [cell.value for cell in ws['A'] if cell.value is not None]
    a = generate_macro('last_month', accounts, 'last_month')
    b = generate_macro('campaigns', accounts, 'campaigns')
    return a, b
