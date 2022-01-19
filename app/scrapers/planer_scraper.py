from bs4 import BeautifulSoup
from openpyxl import Workbook
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as ec

from .agency_driver import AgencyDriver


def generate_url_from_phrase(phrase):
    base = 'https://ads.allegro.pl/panel/planner?phrase='
    phrase = phrase.replace(" ", '%20')
    return base + phrase


class PlanerResult:
    def __init__(self, phrase: str, concurrency: str, avg_monthly_coverage: str, cpc: str):
        self.phrase = phrase
        if concurrency not in ('duża', 'mała', 'średnia'):
            raise ValueError('wrong concurrency type!')
        self.concurrency = concurrency
        self.avg_monthly_coverage = avg_monthly_coverage
        self.cpc = cpc

    def __eq__(self, other):
        return self.phrase == other.phrase and self.concurrency == other.concurrency

    def __repr__(self):
        return self.phrase, self.concurrency, self.avg_monthly_coverage, self.cpc

    def __hash__(self):
        return hash(self.__repr__())

    @property
    def is_small(self) -> bool:
        return self.concurrency == 'mała'

    @property
    def is_medium(self) -> bool:
        return self.concurrency == 'średnia'


def scrape_results(page_source):
    soup = BeautifulSoup(page_source, 'html5lib')
    table_body = soup.find('tbody')
    trs = table_body.find_all('tr')

    results = []

    for tr in trs:
        td = tr.find_all('td')
        phrase = td[0].a.text
        avg_monthly_coverage = td[1].text
        cpc = td[2].text
        concurrency = td[3].text
        results.append(PlanerResult(phrase, concurrency, avg_monthly_coverage, cpc))

    return results


def scrape_phrase(driver: AgencyDriver, phrase: str):
    driver.driver.get(generate_url_from_phrase(phrase))
    try:
        driver.wait(ec.presence_of_element_located((By.CSS_SELECTOR, 'table')), 1)
    except TimeoutException:
        return []
    return scrape_results(driver.page_source)


def sort_results(e: PlanerResult):
    if e.avg_monthly_coverage == '< 200':
        return 199
    else:
        return int(e.avg_monthly_coverage.replace(" ", ""))


def scrape_planner(phrase: str):
    results = set()
    scraped_phrases = set()
    driver = AgencyDriver()

    first_level_results = scrape_phrase(driver, phrase)
    results.update(first_level_results)

    second_level_results = []
    for result in first_level_results:
        scraped_phrases.add(result.phrase)
        second_level_results += (scrape_phrase(driver, result.phrase))

    second_level_results = set(second_level_results)
    results.update(second_level_results)

    third_level_results = []
    for result in second_level_results:
        if result.phrase not in scraped_phrases:
            third_level_results += scrape_phrase(driver, result.phrase)

    results.update(third_level_results)

    wb = Workbook()
    ws = wb.active
    row = 1
    ws.cell(row=row, column=1).value = 'mała konkurencja'
    ws.cell(row=row, column=2).value = 'fraza'
    ws.cell(row=row, column=3).value = 'śr. miesięczny zasięg'
    ws.cell(row=row, column=4).value = 'cpc'

    row += 1
    small = [result for result in results if result.is_small]
    small.sort(reverse=True, key=sort_results)
    for r in small:
        ws.cell(row=row, column=2).value = r.phrase
        ws.cell(row=row, column=3).value = r.avg_monthly_coverage
        ws.cell(row=row, column=4).value = r.cpc
        row += 1

    row = 1

    ws.cell(row=row, column=6).value = 'średnia konkurencja'
    ws.cell(row=row, column=7).value = 'fraza'
    ws.cell(row=row, column=8).value = 'śr. miesięczny zasięg'
    ws.cell(row=row, column=9).value = 'cpc'

    row += 1
    medium = [result for result in results if result.is_medium]
    medium.sort(reverse=True, key=sort_results)
    for r in medium:
        ws.cell(row=row, column=7).value = r.phrase
        ws.cell(row=row, column=8).value = r.avg_monthly_coverage
        ws.cell(row=row, column=9).value = r.cpc
        row += 1

    wb.save(f'/home/kajetan/Documents/pryzmat/planer_scraper/{phrase}.xlsx')
